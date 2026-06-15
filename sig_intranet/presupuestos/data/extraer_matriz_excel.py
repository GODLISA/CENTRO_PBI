"""
Extrae la matriz de cobros del Excel "MATRIZ DE CALCULO - Cobros a PI.xlsx"
y genera matriz_inicial.json (insumo del comando `manage.py cargar_matriz`).

Uso (si llega una versión nueva del Excel):

    python sig_intranet/presupuestos/data/extraer_matriz_excel.py "ruta\\al\\excel.xlsx"

Luego ejecutar `python manage.py cargar_matriz` para actualizar la BD.
Requiere openpyxl.
"""
import json
import re
import sys
from decimal import Decimal

import openpyxl

RUTA = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\User_\Downloads\MATRIZ DE CALCULO -Cobros a PI.xlsx'
UF_IVA = Decimal('40820.31')  # valor "UF" base con que la matriz Excel calculó sus montos

wb = openpyxl.load_workbook(RUTA, data_only=True)
salida = {'zonas': [], 'areas': []}


def limpiar(texto):
    return re.sub(r'\s+', ' ', str(texto)).strip()


def uf_desde_clp(valor):
    """Convierte un monto CLP de la matriz a unidades UF+IVA (ej: 81640.62 -> 2)."""
    return float(round(Decimal(str(valor)) / UF_IVA, 2))


# ---------- Zonas y comunas ----------
def extraer_comunas(hoja, fila_ini, col_comuna, col_region, col_km):
    ws = wb[hoja]
    comunas = []
    for fila in ws.iter_rows(min_row=fila_ini, max_row=ws.max_row):
        comuna = fila[col_comuna - 1].value
        region = fila[col_region - 1].value
        if not comuna or not str(comuna).strip():
            continue
        km = fila[col_km - 1].value
        try:
            km = float(km)
        except (TypeError, ValueError):
            km = 0
        comunas.append({
            'nombre': limpiar(comuna),
            'region': limpiar(region) if region else '',
            'km_ida_regreso': km,
        })
    return comunas


salida['zonas'] = [
    {
        'nombre': 'Centro',
        'descripcion': 'Región Metropolitana y V Región (hoja CENTRO-V de la matriz)',
        'comunas': extraer_comunas('CENTRO-V', 3, 1, 2, 7),
    },
    {
        'nombre': 'Norte',
        'descripcion': 'Regiones de Arica a Coquimbo (hoja NORTE de la matriz)',
        'comunas': extraer_comunas('NORTE', 3, 2, 3, 9),
    },
    {
        'nombre': 'Sur',
        'descripcion': "Regiones de O'Higgins a Magallanes (hoja SUR de la matriz)",
        'comunas': extraer_comunas('SUR', 3, 1, 2, 8),
    },
]

# ---------- Ítems comunes de viaje/estadía (hojas NORTE/SUR/Taller) ----------
ITEMS_VIAJE = [
    {'concepto': 'Alojamiento hotel (por técnico)', 'unidad': 'día', 'moneda': 'CLP', 'precio': 65000},
    {'concepto': 'Viático desayuno (por técnico)', 'unidad': 'día', 'moneda': 'CLP', 'precio': 4000},
    {'concepto': 'Viático almuerzo (por técnico)', 'unidad': 'día', 'moneda': 'CLP', 'precio': 10000},
    {'concepto': 'Viático cena (por técnico)', 'unidad': 'día', 'moneda': 'CLP', 'precio': 10000},
    {'concepto': 'Bono diario técnico', 'unidad': 'día', 'moneda': 'CLP', 'precio': 10000},
    {'concepto': 'Bono trabajo festivo/domingo', 'unidad': 'día', 'moneda': 'CLP', 'precio': 25000},
]


# ---------- Instalaciones (Gastro en UF, Taller en CLP) ----------
def extraer_instalacion(hoja, fila_ini, fila_fin, en_uf):
    """
    Recorre los bloques EQUIPO / INSTALACION / MATERIALES / TOTAL.
    Col B: nombre y modelos del equipo; C: detalle; D: valor unitario.
    """
    ws = wb[hoja]
    items = []
    etiquetas = []   # nombre + modelos acumulados del bloque
    pendientes = []  # ítems del bloque a la espera del nombre completo

    vistos = set()

    def cerrar_bloque():
        nonlocal etiquetas, pendientes
        nombre = ' '.join(etiquetas).strip()
        nombre = re.sub(r'\s+', ' ', nombre)
        for it in pendientes:
            concepto = it['concepto'].replace('{EQ}', nombre)[:250]
            # evitar conceptos duplicados entre bloques distintos
            base, n = concepto, 2
            while concepto in vistos:
                concepto = f'{base} ({n})'
                n += 1
            vistos.add(concepto)
            it['concepto'] = concepto
            items.append(it)
        etiquetas, pendientes = [], []

    for fila in ws.iter_rows(min_row=fila_ini, max_row=fila_fin, min_col=2, max_col=6):
        b, c, d = fila[0].value, fila[1].value, fila[2].value
        f_total = fila[4].value
        b = limpiar(b) if b else ''
        c = limpiar(c) if c else ''

        if b and 'OBSERVACION' in b.upper():
            break
        if b:
            etiquetas.append(b)

        cu = c.upper()
        # Fila de total del bloque: "TOTAL + IVA" en col D, o fila vacía con
        # solo el total en col F (algunos bloques no llevan etiqueta).
        if isinstance(d, str) and 'TOTAL' in d.upper():
            if 'SUB' not in d.upper():
                cerrar_bloque()
            continue
        if not b and not c and d is None and isinstance(f_total, (int, float)):
            cerrar_bloque()
            continue
        if not c or d is None or not isinstance(d, (int, float)):
            continue

        if 'TRASLADO' in cu:
            continue  # el traslado lo calcula el sistema (km x valor km)
        if 'PROVISION' in cu:
            continue
        if 'VISITA' in cu:
            pendientes.append({
                'concepto': 'Visita técnica (incluye 1 hr M.O.)',
                'unidad': 'visita', 'moneda': 'UF', 'precio': uf_desde_clp(d),
            })
        elif 'HR. ADICIONAL' in cu or 'HR ADICIONAL' in cu:
            pendientes.append({
                'concepto': 'Hora adicional mano de obra',
                'unidad': 'hora', 'moneda': 'UF', 'precio': uf_desde_clp(d),
            })
        elif 'INSTALACION' in cu:
            extra = re.search(r'\((CONSIDERA[^)]*|\d+\s*D[SI]A?S?[^)]*)\)', c, re.I)
            detalle = f' ({limpiar(extra.group(1)).capitalize()})' if extra else ''
            if en_uf:
                pendientes.append({
                    'concepto': 'Instalación {EQ}' + detalle,
                    'unidad': 'servicio', 'moneda': 'UF', 'precio': uf_desde_clp(d),
                })
            else:
                pendientes.append({
                    'concepto': 'Instalación {EQ}' + detalle,
                    'unidad': 'servicio', 'moneda': 'CLP', 'precio': round(d),
                })
        elif 'MATERIALES' in cu or 'ARRIENDO' in cu:
            etiqueta = 'Arriendo grúa/alza hombre' if 'ARRIENDO' in cu else 'Materiales'
            pendientes.append({
                'concepto': etiqueta + ' {EQ}',
                'unidad': 'servicio', 'moneda': 'CLP', 'precio': round(d),
            })
    cerrar_bloque()
    return items


items_gastro = extraer_instalacion('Instalación Eq. Gastro', 15, 133, en_uf=True)
items_taller = extraer_instalacion('Instalación Eq. Taller (2)', 10, 74, en_uf=False)

# ---------- Mantención preventiva Eq. Gastro ----------
ws = wb['MP Eq.Gastro']
items_mp = []
for fila in ws.iter_rows(min_row=3, max_row=114, min_col=1, max_col=6):
    marca, equipo, modelo, frecuencia, valor_uf, hrs = (c.value for c in fila)
    if not marca or valor_uf is None or not isinstance(valor_uf, (int, float)):
        continue
    concepto = f'MP {limpiar(marca)} {limpiar(equipo)}'
    if modelo:
        concepto += f' ({limpiar(modelo)})'
    if frecuencia and limpiar(frecuencia).upper() != 'N/A':
        concepto += f' - frecuencia {limpiar(frecuencia).lower()}'
    items_mp.append({
        'concepto': concepto[:255], 'unidad': 'mantención',
        'moneda': 'UF', 'precio': float(valor_uf),
    })

salida['areas'] = [
    {
        'nombre': 'Instalación Equipos Gastronómicos',
        'descripcion': 'Matriz hoja "Instalación Eq. Gastro". Instalación en horas M.O. (1 UF/hr) + materiales.',
        'items': items_gastro + ITEMS_VIAJE,
    },
    {
        'nombre': 'Instalación Equipos de Taller',
        'descripcion': 'Matriz hoja "Instalación Eq. Taller". Valores CLP por equipo (M.O. taller: 0,85 UF/hr).',
        'items': items_taller + ITEMS_VIAJE + [
            {'concepto': 'Viático diario por persona (taller)', 'unidad': 'día', 'moneda': 'CLP', 'precio': 25000},
            {'concepto': 'Bono viaje (por semana)', 'unidad': 'semana', 'moneda': 'CLP', 'precio': 40000},
        ],
    },
    {
        'nombre': 'Mantención Preventiva Equipos Gastronómicos',
        'descripcion': 'Matriz hoja "MP Eq.Gastro". Valor por mantención en UF (UF del día + IVA), no incluye visita ni repuestos.',
        'items': items_mp + [
            {'concepto': 'Visita técnica (incluye 1 hr M.O.)', 'unidad': 'visita', 'moneda': 'UF', 'precio': 2},
            {'concepto': 'Hora adicional mano de obra', 'unidad': 'hora', 'moneda': 'UF', 'precio': 1},
        ] + ITEMS_VIAJE,
    },
]

import os
destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matriz_inicial.json')
with open(destino, 'w', encoding='utf-8') as f:
    json.dump(salida, f, ensure_ascii=False, indent=1)

print('zonas:', [(z['nombre'], len(z['comunas'])) for z in salida['zonas']])
print('areas:', [(a['nombre'], len(a['items'])) for a in salida['areas']])
